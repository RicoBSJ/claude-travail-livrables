#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Met à jour la colonne « Numéro de la dernière leçon » du tableau des jobs.

Le numéro n'est pas compté : il est LU dans le nom du dernier fichier produit
(`YYYY-MM-DD_lecon-<parcours>_NN_slug.docx`), puis on retient le maximum. Compter
les fichiers donnerait un résultat faux au premier trou de numérotation — et un
créneau manqué en crée un (cf. astrologie-karmique le 27/08/2026).

Le lien job → parcours vient de `jobs_config.json`, source de vérité unique : le
slug est extrait du champ `livrable`. Aucune correspondance codée en dur, donc
rien à maintenir ici quand un parcours est ajouté ou supprimé.

Appelé par run_job.sh après chaque job réussi. Idempotent : n'écrit le fichier
que si une valeur a réellement changé.

Usage : python3 outils/scripts/maj_xlsx_jobs.py [--verbeux]
"""
import json
import re
import sys
from copy import copy
from pathlib import Path

RACINE = Path(__file__).resolve().parent.parent.parent
CONFIG = RACINE / "jobs_config.json"
LECONS = RACINE / "livrables" / "lecons"
CLASSEUR = RACINE / "livrables" / "documents" / "2026-08-09_Les 14 Jobs.xlsx"

EN_TETE = "Numéro de la dernière leçon"
SANS_OBJET = "—"          # jobs qui ne produisent pas de leçon (veilles, contrôle)
AUCUNE = "0"              # parcours déclaré mais sans leçon encore produite

verbeux = "--verbeux" in sys.argv


def parcours_par_job():
    """{id_du_job: slug_de_parcours} pour les seuls jobs produisant des leçons."""
    donnees = json.loads(CONFIG.read_text(encoding="utf-8"))
    jobs = donnees["jobs"] if isinstance(donnees, dict) and "jobs" in donnees else donnees
    table = {}
    for job in jobs:
        trouve = re.search(r"lecon-([a-z0-9-]+)_NN", job.get("livrable", ""))
        if trouve:
            table[job["id"]] = trouve.group(1)
    return table


def dernier_numero(parcours):
    """Plus grand NN trouvé dans les .docx du parcours, ou None si aucun.

    Le motif exige le séparateur `_` après le slug : sans lui, « appli-ia »
    capturerait un hypothétique « appli-ia-avance ».
    """
    motif = re.compile(r"_lecon-" + re.escape(parcours) + r"_(\d+)_")
    numeros = [
        int(m.group(1))
        for f in LECONS.glob("*.docx")
        if not f.name.startswith("~$") and (m := motif.search(f.name))
    ]
    return max(numeros) if numeros else None


def main():
    try:
        import openpyxl
    except ImportError:
        print("⚠️ openpyxl absent — colonne non mise à jour (pip3 install openpyxl)")
        return 0
    if not CLASSEUR.exists():
        print(f"⚠️ Classeur introuvable : {CLASSEUR.name} — rien à faire")
        return 0

    table = parcours_par_job()
    classeur = openpyxl.load_workbook(CLASSEUR)
    feuille = classeur.active

    # Repérer la colonne « Job » et la ligne d'en-tête, plutôt que de les supposer.
    ligne_entete = col_job = None
    for ligne in feuille.iter_rows(min_row=1, max_row=6):
        for cellule in ligne:
            if isinstance(cellule.value, str) and cellule.value.strip() == "Job":
                ligne_entete, col_job = cellule.row, cellule.column
                break
        if ligne_entete:
            break
    if not ligne_entete:
        print("⚠️ Colonne « Job » introuvable — structure inattendue, rien n'est modifié")
        return 1

    # Colonne cible : celle qui porte déjà l'en-tête, sinon la première libre.
    col_cible = None
    for cellule in feuille[ligne_entete]:
        if isinstance(cellule.value, str) and cellule.value.strip() == EN_TETE:
            col_cible = cellule.column
    modifie = col_cible is None          # vrai uniquement à la création de la colonne
    if modifie:
        col_cible = feuille.max_column + 1
        modele = feuille.cell(row=ligne_entete, column=feuille.max_column)
        cible = feuille.cell(row=ligne_entete, column=col_cible, value=EN_TETE)
        cible._style = copy(modele._style)
        feuille.column_dimensions[cible.column_letter].width = 26
        if verbeux:
            print(f"colonne créée : {cible.column_letter} — « {EN_TETE} »")

    lignes = 0

    for ligne in range(ligne_entete + 1, feuille.max_row + 1):
        job = feuille.cell(row=ligne, column=col_job).value
        if not isinstance(job, str) or not job.strip():
            continue
        job = job.strip()
        if job in table:
            numero = dernier_numero(table[job])
            valeur = str(numero) if numero is not None else AUCUNE
        else:
            valeur = SANS_OBJET

        cellule = feuille.cell(row=ligne, column=col_cible)
        if cellule.value != valeur:
            cellule.value = valeur
            modifie = True
            if verbeux:
                print(f"  {job:28} → {valeur}")
        # aligner sur le style de la colonne précédente, sans écraser le contenu
        cellule._style = copy(feuille.cell(row=ligne, column=col_cible - 1)._style)
        lignes += 1

    if modifie:
        classeur.save(CLASSEUR)
        print(f"[xlsx] Colonne « {EN_TETE} » mise à jour ({lignes} lignes).")
    elif verbeux:
        print(f"[xlsx] Aucun changement ({lignes} lignes vérifiées).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
